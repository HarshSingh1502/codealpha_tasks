"""Export scraped records to JSON, CSV, and Excel."""

from __future__ import annotations

import csv
import json
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _fieldnames(records: list[dict[str, Any]]) -> list[str]:
    names: list[str] = []
    for record in records:
        for key in record:
            if key not in names:
                names.append(key)
    return names


def export_json(records: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(records, handle, indent=2, ensure_ascii=False)


def export_csv(records: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not records:
        path.write_text("", encoding="utf-8")
        return

    fieldnames = _fieldnames(records)

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def export_excel(records: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(export_excel_bytes(records))


def export_excel_bytes(records: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scraped Items"

    if not records:
        sheet.append(["message"])
        sheet.append(["No items found"])
    else:
        columns = _fieldnames(records)
        sheet.append(columns)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for record in records:
            sheet.append([record.get(col, "") for col in columns])

        for index, column in enumerate(columns, start=1):
            letter = get_column_letter(index)
            max_len = len(column)
            for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
                value = row[0].value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            sheet.column_dimensions[letter].width = min(max_len + 2, 60)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
